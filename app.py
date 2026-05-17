import io
import os
import uuid
from datetime import datetime, timezone

import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from db_manager import db
from invoice_parser import parse_invoice, categorize

load_dotenv()


def _secret(key: str, default: str = "") -> str:
    """Read from st.secrets (Streamlit Cloud) with os.environ as local fallback."""
    try:
        return st.secrets[key]
    except Exception:
        return os.environ.get(key, default)


st.set_page_config(page_title="Vessel Provisioning System", layout="wide")

CATEGORIES = ['Chilled', 'Dry', 'Frozen', 'Softdrink']
ALL_SUB_CATS = ['Beverages', 'Bread & Bakery', 'Cans & Preserves', 'Dairy',
                'Grains & Baking', 'Meat & Deli', 'Meat & Seafood',
                'Oils & Spices', 'Vegetables', 'Vegetables & Fruit', 'Other']


# ── PIN Authentication ─────────────────────────────────────────────────────────
if not st.session_state.get('authenticated'):
    _, col, _ = st.columns([1, 1, 1])
    with col:
        st.title("Vessel Provisioning System")
        st.caption("ESVAGT DANA")
        st.divider()
        pin = st.text_input("PIN Code", type="password", key="pin_input",
                            placeholder="Enter PIN")
        if st.button("Login", type="primary", use_container_width=True):
            if pin == _secret("APP_PIN"):
                st.session_state['authenticated'] = True
                st.rerun()
            else:
                st.error("Incorrect PIN. Please try again.")
    st.stop()

# ── Scoped CSS — strictly targets column rows that contain a number input ───────
# :has(stNumberInput) means ONLY item-card rows are affected.
# Button rows, nav bar, and dashboard grids never match this selector.
st.markdown("""
<style>
/* Force item-card rows to stay horizontal on all screen sizes */
div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"]) {
    flex-wrap: nowrap !important;
    align-items: flex-start !important;
    gap: 0.25rem !important;
    padding-bottom: 4px !important;
    margin-bottom: 0 !important;
}
div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"])
    > div[data-testid="stColumn"] {
    min-width: 0 !important;
    padding-top: 2px !important;
    padding-bottom: 2px !important;
}
/* Collapse paragraph margins inside item-card info column */
div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"])
    .stMarkdown p {
    margin: 0 !important;
    line-height: 1.25 !important;
}
/* Remove top gap above number input inside item-card rows */
div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"])
    div[data-testid="stNumberInput"] {
    margin-top: 0 !important;
    padding-top: 0 !important;
}
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=300)
def load_inventory():
    docs = db.collection('inventory').stream()
    items = [{'id': doc.id, **doc.to_dict()} for doc in docs]
    items.sort(key=lambda x: (x.get('category', ''), x.get('sub_category', ''), x.get('name', '')))
    return items


def load_counts():
    docs = db.collection('counts').stream()
    counts = [{'id': doc.id, **doc.to_dict()} for doc in docs]
    counts.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return counts


def make_excel_from_rows(rows):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Proviant Status"
    if not rows:
        wb.save(output)
        return output.getvalue()
    df = pd.DataFrame(rows)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    col_widths = {
        'Category': 12, 'Sub-Category': 22, 'Item Name': 52,
        'Unit': 8, 'Price (DKK)': 14, 'Quantity': 10, 'Total Value (DKK)': 18,
    }
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 16)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Grand Total row
    total_row_idx = len(df) + 2
    grand_total = sum(r.get('Total Value (DKK)', 0) for r in rows)
    columns = list(df.columns)
    item_col  = columns.index('Item Name') + 1
    total_col = columns.index('Total Value (DKK)') + 1
    total_font = Font(bold=True)
    label_cell = ws.cell(row=total_row_idx, column=item_col, value="Grand Total")
    label_cell.font = total_font
    value_cell = ws.cell(row=total_row_idx, column=total_col, value=round(grand_total, 2))
    value_cell.font = total_font

    wb.save(output)
    return output.getvalue()


def auto_save_qty(item_id: str, count_id: str) -> None:
    """Persist a single item's quantity to counts/{count_id}.saved_quantities."""
    qty = int(st.session_state.get(f"qty_{count_id}_{item_id}", 0))
    print(f"[SAVE] count={count_id[:8]}... item={item_id!r} qty={qty}")
    cache_key = f'_qtys_cache_{count_id}'
    saved = dict(st.session_state.get(cache_key, {}))
    saved[item_id] = qty
    st.session_state[cache_key] = saved
    try:
        db.collection('counts').document(count_id).update({'saved_quantities': saved})
        print(f"[SAVE] OK — {len(saved)} item(s) in map")
    except Exception as exc:
        print(f"[SAVE] ERROR: {exc}")


def render_item_card(item, count_id: str = None, show_category: bool = False):
    """Compact side-by-side card: name+price stacked in left col, qty input in right col."""
    unit  = item.get('unit', '')
    price = item.get('price', 0)
    price_str = f"{price:.2f} DKK / {unit}" if unit else f"{price:.2f} DKK"
    key = f"qty_{count_id}_{item['id']}" if count_id else f"qty_{item['id']}"

    if show_category:
        st.caption(f"{item.get('category', '')}  ·  {item.get('sub_category', '')}")
    col_info, col_input = st.columns([3, 1])
    with col_info:
        st.markdown(f"**{item.get('name', '')}**")
        st.caption(price_str)
    with col_input:
        on_change_kwargs = (
            {'on_change': auto_save_qty, 'args': (item['id'], count_id)}
            if count_id else {}
        )
        st.number_input(
            label=item.get('name', ''),
            min_value=0,
            step=1,
            value=st.session_state.get(key, 0),
            key=key,
            label_visibility="collapsed",
            **on_change_kwargs,
        )
    st.markdown(
        "<div style='border-top:1px solid #e8e4dc;margin:2px 0 0 0;padding:0;font-size:0;line-height:0;'></div>",
        unsafe_allow_html=True,
    )


# ── Screen routing ─────────────────────────────────────────────────────────────
if 'screen' not in st.session_state:
    st.session_state['screen'] = 'home'


# ── SCREEN 2: Active Count ─────────────────────────────────────────────────────
if st.session_state['screen'] == 'count':
    count_id   = st.session_state.get('active_count_id', '')
    count_name = st.session_state.get('active_count_name', 'Count')

    # Restore saved quantities on first entry into this count
    if not st.session_state.get(f'_loaded_{count_id}'):
        print(f"[LOAD] Fetching saved quantities for count: {count_id}")
        all_items = load_inventory()
        for item in all_items:
            st.session_state[f"qty_{count_id}_{item['id']}"] = 0
        count_doc = db.collection('counts').document(count_id).get()
        saved: dict = {}
        if count_doc.exists:
            saved = count_doc.to_dict().get('saved_quantities', {})
            for item_id, qty in saved.items():
                st.session_state[f"qty_{count_id}_{item_id}"] = int(qty)
        st.session_state[f'_qtys_cache_{count_id}'] = dict(saved)
        st.session_state[f'_loaded_{count_id}'] = True
        print(f"[LOAD] Loaded {len(saved)} saved quantities")

    # Top navigation bar
    col_back, col_title, col_finish = st.columns([2, 6, 2])
    with col_back:
        if st.button("Back to Dashboard", use_container_width=True):
            all_items = load_inventory()
            quantities = {
                item['id']: st.session_state.get(f"qty_{count_id}_{item['id']}", 0)
                for item in all_items
            }
            quantities = {k: v for k, v in quantities.items() if v > 0}
            db.collection('counts').document(count_id).update({
                'saved_quantities': quantities
            })
            st.session_state[f'_loaded_{count_id}'] = False
            st.session_state['screen'] = 'home'
            st.rerun()
    with col_title:
        st.title(count_name)
    with col_finish:
        if st.button("Finish Count", type="primary", use_container_width=True):
            all_items = load_inventory()
            rows = []
            for item in all_items:
                qty = st.session_state.get(f"qty_{count_id}_{item['id']}", 0)
                if qty > 0:
                    rows.append({
                        'Category':          item.get('category', ''),
                        'Sub-Category':      item.get('sub_category', 'Other'),
                        'Item Name':         item.get('name', ''),
                        'Unit':              item.get('unit', ''),
                        'Price (DKK)':       round(item.get('price', 0.0), 2),
                        'Quantity':          int(qty),
                        'Total Value (DKK)': round(item.get('price', 0.0) * qty, 2),
                    })
            total_value = sum(r['Total Value (DKK)'] for r in rows)
            db.collection('counts').document(count_id).update({
                'status':       'completed',
                'items':        rows,
                'total_value':  total_value,
                'completed_at': datetime.now(timezone.utc).isoformat(),
            })
            st.session_state['screen'] = 'home'
            st.rerun()

    st.divider()

    # Search bar
    search_query = st.text_input(
        "Search items", placeholder="Type to filter items...",
        key="count_search", label_visibility="collapsed",
    )

    items = load_inventory()

    if search_query.strip():
        # ── Flat search results view ───────────────────────────────────────────
        filtered = [i for i in items
                    if search_query.lower() in i.get('name', '').lower()]
        if filtered:
            st.caption(f"{len(filtered)} item(s) found")
            for item in filtered:
                render_item_card(item, count_id=count_id, show_category=True)
        else:
            st.info("No items match your search.")
    else:
        # ── Nested expander view ───────────────────────────────────────────────
        cat_list = sorted(set(item.get('category', '') for item in items))
        for cat in cat_list:
            cat_items = [i for i in items if i.get('category') == cat]
            with st.expander(f"{cat}  —  {len(cat_items)} items", expanded=False):
                sub_cats = sorted(set(i.get('sub_category', 'Other') for i in cat_items))
                for sub_cat in sub_cats:
                    sub_items = [i for i in cat_items if i.get('sub_category', 'Other') == sub_cat]
                    with st.expander(f"{sub_cat}  ({len(sub_items)})", expanded=False):
                        for item in sub_items:
                            render_item_card(item, count_id=count_id)

    # Live preview
    st.divider()
    st.subheader("Live Preview")
    preview_rows = []
    for item in items:
        qty = st.session_state.get(f"qty_{count_id}_{item['id']}", 0)
        if qty > 0:
            preview_rows.append({
                'Category':          item.get('category', ''),
                'Sub-Category':      item.get('sub_category', 'Other'),
                'Item Name':         item.get('name', ''),
                'Unit':              item.get('unit', ''),
                'Price (DKK)':       round(item.get('price', 0.0), 2),
                'Quantity':          int(qty),
                'Total Value (DKK)': round(item.get('price', 0.0) * qty, 2),
            })
    if preview_rows:
        total = sum(r['Total Value (DKK)'] for r in preview_rows)
        st.metric("Total Inventory Value", f"{total:,.2f} DKK")
        df = (pd.DataFrame(preview_rows)
                .sort_values(['Category', 'Sub-Category', 'Item Name'])
                .reset_index(drop=True))
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("No items with quantity > 0 yet.")

    st.stop()


# ── SCREEN 1: Home ─────────────────────────────────────────────────────────────
st.title("Vessel Provisioning System")
st.caption("ESVAGT DANA — Inventory overview")

# Create new count
with st.container(border=True):
    st.subheader("Create New Count")
    col_inp, col_btn = st.columns([4, 1])
    with col_inp:
        new_name = st.text_input(
            "Name", placeholder="e.g. May 2025 Count",
            key="new_count_name", label_visibility="collapsed",
        )
    with col_btn:
        if st.button("Create", type="primary", use_container_width=True):
            if new_name.strip():
                count_id = str(uuid.uuid4())
                db.collection('counts').document(count_id).set({
                    'name':             new_name.strip(),
                    'status':           'ongoing',
                    'created_at':       datetime.now(timezone.utc).isoformat(),
                    'saved_quantities': {},
                    'items':            [],
                    'total_value':      0.0,
                })
                st.session_state['screen']            = 'count'
                st.session_state['active_count_id']   = count_id
                st.session_state['active_count_name'] = new_name.strip()
                st.rerun()
            else:
                st.warning("Please enter a name for the count.")

st.divider()

counts    = load_counts()
ongoing   = [c for c in counts if c.get('status') == 'ongoing']
completed = [c for c in counts if c.get('status') == 'completed']

# Active counts
st.subheader(f"Active Counts  ({len(ongoing)})")
if ongoing:
    cols = st.columns(3)
    for idx, count in enumerate(ongoing):
        cid = count['id']
        with cols[idx % 3]:
            with st.container(border=True):
                if st.session_state.get(f'_renaming_{cid}'):
                    # ── Rename mode ────────────────────────────────────────────
                    new_title = st.text_input(
                        "New name", value=count.get('name', ''),
                        key=f'_rename_input_{cid}', label_visibility='collapsed',
                    )
                    sv_col, cx_col = st.columns(2)
                    with sv_col:
                        if st.button("Save", key=f'_save_rename_{cid}',
                                     use_container_width=True, type='primary'):
                            if new_title.strip():
                                db.collection('counts').document(cid).update(
                                    {'name': new_title.strip()}
                                )
                            st.session_state[f'_renaming_{cid}'] = False
                            st.rerun()
                    with cx_col:
                        if st.button("Cancel", key=f'_cancel_rename_{cid}',
                                     use_container_width=True):
                            st.session_state[f'_renaming_{cid}'] = False
                            st.rerun()

                elif st.session_state.get(f'_confirm_del_{cid}'):
                    # ── Delete-confirm mode ────────────────────────────────────
                    st.warning(f"Delete **{count.get('name', 'Unnamed')}**?")
                    yes_col, cx_col = st.columns(2)
                    with yes_col:
                        if st.button("Yes, Delete", key=f'_del_yes_{cid}',
                                     use_container_width=True, type='primary'):
                            db.collection('counts').document(cid).delete()
                            st.session_state[f'_confirm_del_{cid}'] = False
                            st.rerun()
                    with cx_col:
                        if st.button("Cancel", key=f'_cancel_del_{cid}',
                                     use_container_width=True):
                            st.session_state[f'_confirm_del_{cid}'] = False
                            st.rerun()

                else:
                    # ── Normal mode ────────────────────────────────────────────
                    st.markdown(f"**{count.get('name', 'Unnamed')}**")
                    st.caption(f"Created: {count.get('created_at', '')[:10]}")
                    if st.button("Open", key=f"open_{cid}", use_container_width=True):
                        st.session_state['screen']            = 'count'
                        st.session_state['active_count_id']   = cid
                        st.session_state['active_count_name'] = count.get('name', 'Count')
                        st.rerun()
                    ren_col, del_col = st.columns(2)
                    with ren_col:
                        if st.button("Rename", key=f'_rename_{cid}',
                                     use_container_width=True):
                            st.session_state[f'_renaming_{cid}'] = True
                            st.rerun()
                    with del_col:
                        if st.button("Delete", key=f'_del_{cid}',
                                     use_container_width=True):
                            st.session_state[f'_confirm_del_{cid}'] = True
                            st.rerun()
else:
    st.info("No active counts.")

st.divider()

# Completed counts
st.subheader(f"Completed Counts  ({len(completed)})")
if completed:
    cols = st.columns(3)
    for idx, count in enumerate(completed):
        with cols[idx % 3]:
            with st.container(border=True):
                st.markdown(f"**{count.get('name', 'Unnamed')}**")
                completed_date = count.get('completed_at', count.get('created_at', ''))
                st.caption(f"Completed: {completed_date[:10]}")
                total = count.get('total_value', 0.0)
                st.write(f"Total: {total:,.2f} DKK")
                dl_col, del_col = st.columns(2)
                with dl_col:
                    excel_data = make_excel_from_rows(count.get('items', []))
                    st.download_button(
                        label="Download",
                        data=excel_data,
                        file_name=f"{count.get('name', 'count').replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{count['id']}",
                        use_container_width=True,
                    )
                with del_col:
                    if st.button("Delete", key=f"del_{count['id']}", use_container_width=True):
                        db.collection('counts').document(count['id']).delete()
                        st.rerun()
else:
    st.info("No completed counts.")


# ── Admin tools ────────────────────────────────────────────────────────────────
st.divider()
with st.expander("Update Prices (Upload PDF)", expanded=False):
    st.write(
        "Upload a new invoice PDF to refresh prices and units. "
        "Stock quantities and manually-set categories are preserved. "
        "Items not yet in the database are added automatically."
    )

    if 'upload_msg' in st.session_state:
        st.success(st.session_state.pop('upload_msg'))

    uploaded = st.file_uploader("Select invoice PDF", type="pdf", key="pdf_uploader")

    if uploaded and st.button("Process and Update Prices", type="primary"):
        with st.spinner("Parsing PDF..."):
            parsed_items = parse_invoice(io.BytesIO(uploaded.getvalue()))

        with st.spinner(f"Updating {len(parsed_items)} items in Firestore..."):
            existing_ids = {doc.id for doc in db.collection('inventory').stream()}
            updated = added = 0
            for item in parsed_items:
                doc_id  = item['desc'].replace('/', ' ')
                price   = item['amount'] if item['amount'] is not None else 0.0
                unit    = item.get('unit', '')
                sub_cat = item.get('sub_category', 'Other')
                if doc_id in existing_ids:
                    # Existing item: update price only — preserve manual
                    # category, sub_category, and unit overrides.
                    db.collection('inventory').document(doc_id).update({
                        'name': item['desc'],
                        'price': price,
                    })
                    updated += 1
                else:
                    db.collection('inventory').document(doc_id).set({
                        'name': item['desc'], 'price': price,
                        'unit': unit, 'sub_category': sub_cat,
                        'category': categorize(item['desc']), 'stock': 0,
                    })
                    added += 1

        load_inventory.clear()
        st.session_state['upload_msg'] = (
            f"Done! Updated {updated} existing items, added {added} new."
        )
        st.rerun()


with st.expander("Admin — Fix Categories", expanded=False):
    st.write("Changes are saved to Firestore immediately.")

    search = st.text_input("Filter items", placeholder="Type to search by name...",
                           key="admin_search")

    def on_category_change(item_id):
        db.collection('inventory').document(item_id).update(
            {'category': st.session_state[f"cat_{item_id}"]}
        )
        load_inventory.clear()

    def on_sub_category_change(item_id):
        db.collection('inventory').document(item_id).update(
            {'sub_category': st.session_state[f"sub_{item_id}"]}
        )
        load_inventory.clear()

    admin_items = load_inventory()
    if search:
        admin_items = [i for i in admin_items
                       if search.lower() in i.get('name', '').lower()]

    st.caption(f"Showing {len(admin_items)} items")
    hdr = st.columns([4, 2, 2, 1, 1])
    hdr[0].markdown("**Item**")
    hdr[1].markdown("**Category**")
    hdr[2].markdown("**Sub-Category**")

    for item in admin_items:
        item_id     = item['id']
        current_cat = item.get('category', 'Dry')
        current_sub = item.get('sub_category', 'Other')

        if st.session_state.get(f'_admin_editing_{item_id}'):
            # ── Inline rename row ──────────────────────────────────────────────
            ecol, bcol = st.columns([5, 3])
            with ecol:
                new_name = st.text_input(
                    "Name", value=item.get('name', ''),
                    key=f'_edit_name_{item_id}', label_visibility='collapsed',
                )
            with bcol:
                sv, cx = st.columns(2)
                with sv:
                    if st.button("Save", key=f'_save_name_{item_id}',
                                 type='primary', use_container_width=True):
                        if new_name.strip():
                            db.collection('inventory').document(item_id).update(
                                {'name': new_name.strip()}
                            )
                            load_inventory.clear()
                        st.session_state[f'_admin_editing_{item_id}'] = False
                        st.rerun()
                with cx:
                    if st.button("Cancel", key=f'_cancel_name_{item_id}',
                                 use_container_width=True):
                        st.session_state[f'_admin_editing_{item_id}'] = False
                        st.rerun()

        elif st.session_state.get(f'_admin_confirm_del_{item_id}'):
            # ── Delete-confirm row ─────────────────────────────────────────────
            dcol, bcol = st.columns([5, 3])
            with dcol:
                st.warning(f"Delete **{item.get('name', '')}**?")
            with bcol:
                yes, cx = st.columns(2)
                with yes:
                    if st.button("Yes, Delete", key=f'_del_yes_{item_id}',
                                 type='primary', use_container_width=True):
                        db.collection('inventory').document(item_id).delete()
                        load_inventory.clear()
                        st.session_state[f'_admin_confirm_del_{item_id}'] = False
                        st.rerun()
                with cx:
                    if st.button("Cancel", key=f'_cancel_del_{item_id}',
                                 use_container_width=True):
                        st.session_state[f'_admin_confirm_del_{item_id}'] = False
                        st.rerun()

        else:
            # ── Normal row ─────────────────────────────────────────────────────
            row = st.columns([4, 2, 2, 1, 1])
            row[0].write(item.get('name', ''))
            row[1].selectbox(
                label=f"cat_{item_id}", options=CATEGORIES,
                index=CATEGORIES.index(current_cat) if current_cat in CATEGORIES else 0,
                key=f"cat_{item_id}", on_change=on_category_change, args=(item_id,),
                label_visibility="collapsed",
            )
            row[2].selectbox(
                label=f"sub_{item_id}", options=ALL_SUB_CATS,
                index=ALL_SUB_CATS.index(current_sub) if current_sub in ALL_SUB_CATS else 8,
                key=f"sub_{item_id}", on_change=on_sub_category_change, args=(item_id,),
                label_visibility="collapsed",
            )
            with row[3]:
                if st.button("Edit", key=f'_admin_edit_{item_id}',
                             use_container_width=True):
                    st.session_state[f'_admin_editing_{item_id}'] = True
                    st.rerun()
            with row[4]:
                if st.button("Del", key=f'_admin_del_{item_id}',
                             use_container_width=True):
                    st.session_state[f'_admin_confirm_del_{item_id}'] = True
                    st.rerun()
